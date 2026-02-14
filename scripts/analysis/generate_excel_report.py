#!/usr/bin/env python3
"""
ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„±ê¸°

ì¢…í•© ì¶”ì²œ ìˆœìœ„ì™€ ë‹¤ì–‘í•œ ê¸°ì¤€ë³„ ìˆœìœ„ë¥¼ ì—‘ì…€ íŒŒì¼ë¡œ ìƒì„±í•©ë‹ˆë‹¤.
"""

import pandas as pd
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed. Excel formatting will be basic.")


def calculate_combined_score(df: pd.DataFrame, signal_bonus: int = 5) -> pd.DataFrame:
    """
    ì¢…í•© ì ìˆ˜ ê³„ì‚° (Aì•ˆ: ë‹¨ìˆœ ë³´ë„ˆìŠ¤ ë°©ì‹)

    ì¢…í•©ì ìˆ˜ = ì›ë˜ ì ìˆ˜ + (ì‹œê·¸ë„ ê°œìˆ˜ Ã— 5ì )

    Args:
        df: ë°ì´í„°í”„ë ˆì„
        signal_bonus: ì‹œê·¸ë„ 1ê°œë‹¹ ë³´ë„ˆìŠ¤ ì ìˆ˜ (ê¸°ë³¸ 5ì )

    Returns:
        ì¢…í•©ì ìˆ˜ê°€ ì¶”ê°€ëœ ë°ì´í„°í”„ë ˆì„
    """
    df = df.copy()
    df['combined_score'] = df['score'] + (df['signal_count'] * signal_bonus)
    return df


def format_excel_sheet(ws, df, title=None):
    """ì—‘ì…€ ì‹œíŠ¸ í¬ë§·íŒ…"""
    if not HAS_OPENPYXL:
        return

    # í—¤ë” ìŠ¤íƒ€ì¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # ì œëª© ì¶”ê°€ (ìˆëŠ” ê²½ìš°)
    if title:
        ws.insert_rows(1)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{chr(64 + len(df.columns))}1')
        start_row = 2
    else:
        start_row = 1

    # í—¤ë” í¬ë§·íŒ…
    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ì—´ ë„ˆë¹„ ìë™ ì¡°ì • (í•œê¸€ ê³ ë ¤)
    def get_text_width(text):
        """í…ìŠ¤íŠ¸ ë„ˆë¹„ ê³„ì‚° (í•œê¸€ì€ 2ë°°, ì˜ë¬¸ì€ 1ë°°)"""
        if text is None:
            return 0
        width = 0
        for char in str(text):
            # í•œê¸€, í•œì, ì¼ë³¸ì–´ ë“± (ìœ ë‹ˆì½”ë“œ ë²”ìœ„)
            if '\uac00' <= char <= '\ud7a3' or '\u4e00' <= char <= '\u9fff':
                width += 2  # í•œê¸€/í•œìëŠ” 2ë°° ë„ˆë¹„
            else:
                width += 1  # ì˜ë¬¸/ìˆ«ìëŠ” 1ë°° ë„ˆë¹„
        return width

    for col_idx, column in enumerate(ws.columns, 1):
        max_width = 0
        column_letter = chr(64 + col_idx)  # A, B, C, ...
        for cell in column:
            try:
                # ì œëª© í–‰(ë³‘í•©ëœ ì…€) ìŠ¤í‚µ - í—¤ë” í–‰ë¶€í„° ê³„ì‚°
                if title and cell.row < start_row:
                    continue

                # MergedCell ìŠ¤í‚µ
                if hasattr(cell, 'value') and cell.value is not None:
                    cell_width = get_text_width(cell.value)
                    if cell_width > max_width:
                        max_width = cell_width
            except:
                pass
        # ì—¬ìœ  ê³µê°„ ì¶”ê°€ (+3) ë° ìµœëŒ€ ë„ˆë¹„ ì œí•œ (70)
        adjusted_width = min(max_width + 3, 70) if max_width > 0 else 12
        ws.column_dimensions[column_letter].width = adjusted_width

    # ìˆ«ì í¬ë§·íŒ… ë° ì •ë ¬
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row+1), start=start_row+1):
        for col_idx, cell in enumerate(row, start=1):
            if isinstance(cell.value, (int, float)):
                # ì˜¤ë¥¸ìª½ ì •ë ¬
                cell.alignment = Alignment(horizontal='right')

                # ì»¬ëŸ¼ëª… ê°€ì ¸ì˜¤ê¸°
                header_cell = ws.cell(row=start_row, column=col_idx)
                header = str(header_cell.value) if header_cell.value else ''

                # ì •ìˆ˜ë¡œ í‘œì‹œí•  ì»¬ëŸ¼ (ìˆœìœ„, ì‹œê·¸ë„, ì¢…ëª©ìˆ˜ ë“±)
                integer_columns = ['ìˆœìœ„', 'ì‹œê·¸ë„', 'ì¢…ëª©ìˆ˜', 'ì¢…ëª© ìˆ˜']

                if any(int_col in header for int_col in integer_columns):
                    # ì •ìˆ˜ í¬ë§· (ì†Œìˆ˜ì  ì—†ìŒ)
                    cell.number_format = '0'
                else:
                    # ì‹¤ìˆ˜ í¬ë§· (ì†Œìˆ˜ì  ë‘˜ì§¸ìë¦¬)
                    cell.number_format = '0.00'


def create_glossary_sheet(writer, signal_bonus: int):
    """
    ìš©ì–´ ì„¤ëª… ì‹œíŠ¸ ìƒì„± (ê°€ë¡œ í¼ì¹¨ í‘œ í˜•ì‹)

    Args:
        writer: ExcelWriter ê°ì²´
        signal_bonus: ì‹œê·¸ë„ ë³´ë„ˆìŠ¤ ì ìˆ˜
    """
    # ìš©ì–´ ì„¤ëª… ë°ì´í„° ì¤€ë¹„ (ê°€ë¡œë¡œ í¼ì¹œ í˜•ì‹)
    glossary_data = [
        ['ğŸ“Š ìˆ˜ê¸‰ ë ˆì§ ìŠ¤ìºë„ˆ - ìš©ì–´ ì„¤ëª…', '', '', '', ''],
        ['', '', '', '', ''],

        # ë³´ê³ ì„œ ê°œìš”
        ['â–  ë³´ê³ ì„œ ê°œìš”', '', '', '', ''],
        ['ë¶„ì„ ëª©ì ', 'ì™¸êµ­ì¸/ê¸°ê´€ íˆ¬ìì ìˆ˜ê¸‰ íë¦„ì„ ì •ëŸ‰í™”í•˜ì—¬ ë§¤ìˆ˜ ê°•ë„ê°€ ë†’ì€ ì¢…ëª©ì„ ë°œêµ´', '', '', ''],
        ['ë¶„ì„ ê¸°ê°„', '1ì£¼(1W) ~ 2ë…„(2Y)ê¹Œì§€ 6ê°œ ê¸°ê°„ì„ ë¶„ì„í•˜ì—¬ ë‹¨ê¸°/ì¤‘ê¸°/ì¥ê¸° íŠ¸ë Œë“œë¥¼ ì¢…í•©', '', '', ''],
        ['ë°ì´í„° ì¶œì²˜', 'KOSPI200 + KOSDAQ150 ì´ 345ê°œ í•µì‹¬ ì¢…ëª© (2024-01-02 ~ 2026-01-20)', '', '', ''],
        ['', '', '', '', ''],

        # í•µì‹¬ ì§€í‘œ (ê°€ë¡œ ë°°ì¹˜)
        ['â–  í•µì‹¬ ì§€í‘œ', '', '', '', ''],
        ['', 'ğŸ“Œ Sff (Supply Float Factor)', '', 'ğŸ“Œ Z-Score', ''],
        ['ì •ì˜', 'Sff = (ìˆœë§¤ìˆ˜ ê¸ˆì•¡ / ìœ í†µì‹œì´) Ã— 100', '', 'Z-Score = (í˜„ì¬ê°’ - 60ì¼ í‰ê· ) / 60ì¼ í‘œì¤€í¸ì°¨', ''],
        ['ì˜ë¯¸', 'ì‹œê°€ì´ì•¡ ì™œê³¡ ì œê±°, ìœ í†µë¬¼ëŸ‰ ëŒ€ë¹„ ë§¤ìˆ˜ ê°•ë„ ì •ê·œí™”', '', 'ë³€ë™ì„± ë³´ì •í•˜ì—¬ ì´ìƒ ìˆ˜ê¸‰(í‰ì†Œì™€ ë‹¤ë¥¸ ë§¤ìˆ˜/ë§¤ë„) íƒì§€', ''],
        ['í•´ì„', 'ê°’ì´ í´ìˆ˜ë¡ ìœ í†µë¬¼ëŸ‰ ëŒ€ë¹„ ìˆœë§¤ìˆ˜ ê¸ˆì•¡ì´ í¼', '', '|Z| > 2.0: ì´ìƒ ìˆ˜ê¸‰, +ê°’: ê°•í•œ ë§¤ìˆ˜, -ê°’: ê°•í•œ ë§¤ë„', ''],
        ['', '', '', '', ''],

        # íŒ¨í„´ ë¶„ë¥˜ (ê°€ë¡œ 3ì—´ ë°°ì¹˜)
        ['â–  íŒ¨í„´ ë¶„ë¥˜ (3ê°€ì§€ ìœ í˜•)', '', '', '', ''],
        ['', 'ğŸ”¥ ëª¨ë©˜í…€í˜•', 'ğŸ“ˆ ì§€ì†í˜•', 'ğŸ”„ ì „í™˜í˜•', ''],
        ['íŠ¹ì§•', 'ë‹¨ê¸° ëª¨ë©˜í…€ì´ ë§¤ìš° ê°•í•œ ì¢…ëª©', 'ì¥ê¸°ê°„ ì¼ê´€ëœ ìƒìŠ¹ ì¶”ì„¸ ì¢…ëª©', 'ê³¼ê±° ê°•í–ˆìœ¼ë‚˜ ìµœê·¼ ì•½í™” â†’ ì „í™˜ ëŒ€ê¸°', ''],
        ['ì¡°ê±´', '1W-2Y > 1.0 AND (1W+1M)/2 > 0.5', 'ê°€ì¤‘í‰ê·  > 0.8 AND ì–‘ìˆ˜ ê¸°ê°„ > 70%', 'ê°€ì¤‘í‰ê·  > 0.5 AND 1W-2Y < 0', ''],
        ['íˆ¬ì ìŠ¤íƒ€ì¼', 'ë‹¨ê¸° íŠ¸ë ˆì´ë”©, ëŒíŒŒ ë§¤ë§¤', 'ì¤‘ì¥ê¸° ì¶”ì„¸ ì¶”ì¢…, í¬ì§€ì…˜ íŠ¸ë ˆì´ë”©', 'ì €ê°€ ë§¤ìˆ˜ ê¸°íšŒ í¬ì°©, ì—­ì¶”ì„¸ ë§¤ë§¤', ''],
        ['ìœ„í—˜ë„', 'ë†’ìŒ (ë³€ë™ì„± í¼, ì†ì ˆ ì—„ê²© í•„ìš”)', 'ì¤‘ê°„ (ì•ˆì •ì  ìƒìŠ¹, ì¥ê¸° ë³´ìœ  ê°€ëŠ¥)', 'ë†’ìŒ (ì¶”ì„¸ ì „í™˜ ì‹¤íŒ¨ ê°€ëŠ¥ì„±, ì‹ ì¤‘ ì§„ì…)', ''],
        ['', '', '', '', ''],

        # ì‹œê·¸ë„ (ê°€ë¡œ 3ì—´ ë°°ì¹˜)
        ['â–  ì‹œê·¸ë„ (3ê°€ì§€ íƒ€ì´ë° ì§€í‘œ)', '', '', '', ''],
        ['', 'âœ… MA ê³¨ë“ í¬ë¡œìŠ¤', 'âš¡ ìˆ˜ê¸‰ ê°€ì†ë„', 'ğŸ¤ ì™¸ì¸-ê¸°ê´€ ë™ì¡°ìœ¨', ''],
        ['ì •ì˜', 'ì™¸êµ­ì¸ 5ì¼MA > 20ì¼MA ëŒíŒŒ', '(ìµœê·¼5ì¼ í‰ê· ) / (ì§ì „5ì¼ í‰ê· ) > 1.5ë°°', 'ìµœê·¼ 20ì¼ ì¤‘ ë™ì‹œ ë§¤ìˆ˜ ë¹„ìœ¨ > 50%', ''],
        ['ì˜ë¯¸', 'ë‹¨ê¸° ìˆ˜ê¸‰ì´ ì¥ê¸° ì¶”ì„¸ ìƒí–¥ ëŒíŒŒ â†’ ë§¤ìˆ˜', 'ìˆ˜ê¸‰ ê°•ë„ ê¸‰ì¦ â†’ ëª¨ë©˜í…€ ê°€ì†', 'ë‘ íˆ¬ì ì£¼ì²´ ë™ì‹œ ë§¤ìˆ˜ â†’ í™•ì‹ ë„ ë†’ìŒ', ''],
        ['', '', '', '', ''],

        # ì ìˆ˜ ë©”íŠ¸ë¦­ (ê°€ë¡œ 4ì—´ ë°°ì¹˜)
        ['â–  ì ìˆ˜ ë©”íŠ¸ë¦­ (4ê°€ì§€ ì •ë ¬ ê¸°ì¤€)', '', '', '', ''],
        ['', 'ğŸ“ Recent (í˜„ì¬ ê°•ë„)', 'ğŸ“ Momentum (ê°œì„ ë„)', 'ğŸ“ Weighted (ê°€ì¤‘ í‰ê· )', 'ğŸ“ Average (ì¼ê´€ì„±)'],
        ['ê³„ì‚°ì‹', '(1W + 1M) / 2', '1W - 2Y', '1WÃ—0.30 + 1MÃ—0.25 + 3MÃ—0.20 + 6MÃ—0.15 + 1YÃ—0.07 + 2YÃ—0.03', '(1W + 1M + 3M + 6M + 1Y + 2Y) / 6'],
        ['ì˜ë¯¸', 'ìµœê·¼ 1ì£¼~1ê°œì›” ìˆ˜ê¸‰ ê°•ë„ í‰ê· ', 'ë‹¨ê¸° vs ì¥ê¸° ìˆ˜ê¸‰ ê²©ì°¨ â†’ ì „í™˜ì ', 'ìµœê·¼ì— ë†’ì€ ê°€ì¤‘ì¹˜ ë¶€ì—¬í•œ ì¤‘ì¥ê¸° íŠ¸ë Œë“œ', 'ì „ ê¸°ê°„ ë‹¨ìˆœ í‰ê·  â†’ ì¼ê´€ëœ ìˆ˜ê¸‰'],
        ['í™œìš©', 'í˜„ì¬ ì§„í–‰í˜• ë§¤ìˆ˜ì„¸ íŒŒì•…', 'ê³¼ê±° ëŒ€ë¹„ ìˆ˜ê¸‰ ê°œì„  ì—¬ë¶€ íŒë‹¨', 'ì¤‘ì¥ê¸° ì¶”ì„¸ ë°©í–¥ íŒë‹¨', 'ì „ ê¸°ê°„ ê³ ë¥´ê²Œ ê°•í•œ ì¢…ëª© ë°œêµ´'],
        ['', '', '', '', ''],

        # ì¢…í•©ì ìˆ˜ & ì¶”ì²œ ê¸°ì¤€
        ['â–  ì¢…í•©ì ìˆ˜ & ì¶”ì²œ ê¸°ì¤€', '', '', '', ''],
        ['ì¢…í•©ì ìˆ˜', f'íŒ¨í„´ ì ìˆ˜ + (ì‹œê·¸ë„ ê°œìˆ˜ Ã— {signal_bonus}ì )', '', '', ''],
        ['íŒ¨í„´ ì ìˆ˜', 'RecentÃ—25% + MomentumÃ—25% + WeightedÃ—30% + AverageÃ—20% (0~100ì )', '', '', ''],
        ['ì‹œê·¸ë„ ë³´ë„ˆìŠ¤', f'ì‹œê·¸ë„ 1ê°œë‹¹ +{signal_bonus}ì  (ìµœëŒ€ {signal_bonus*3}ì )', '', '', ''],
        ['', '', '', '', ''],
        ['ì¶”ì²œ ë“±ê¸‰', 'â­â­â­ ê°•ë ¥ ì¶”ì²œ', 'â­â­ ì¶”ì²œ', 'â­ ê´€ì‹¬', ''],
        ['ê¸°ì¤€', 'ì¢…í•©ì ìˆ˜ 80+ AND ì‹œê·¸ë„ 2ê°œ ì´ìƒ', 'ì¢…í•©ì ìˆ˜ 70+ AND ì‹œê·¸ë„ 1ê°œ ì´ìƒ', 'ì¢…í•©ì ìˆ˜ 60+ OR ì‹œê·¸ë„ 2ê°œ ì´ìƒ', ''],
        ['', '', '', '', ''],

        # ì§„ì…/ì²­ì‚° ê¸°ì¤€
        ['â–  ì§„ì…/ì²­ì‚° ê¸°ì¤€', '', '', '', ''],
        ['ì§„ì… ì „ëµ', 'ì‹œê·¸ë„ ë°œìƒ ì‹œì ì—ì„œ ë‹¹ì¼ ì¢…ê°€ ë˜ëŠ” ìµì¼ ì‹œì´ˆê°€ ë§¤ìˆ˜', '', '', ''],
        ['ì†ì ˆ ê¸°ì¤€', 'ì§„ì…ê°€ ëŒ€ë¹„ -7% ë„ë‹¬ ì‹œ ë¬´ì¡°ê±´ ì²­ì‚°', '', '', ''],
        ['ëª©í‘œ ìˆ˜ìµë¥ ', '+15% ë‹¬ì„± ì‹œ 50% ìµì ˆ, +25% ë‹¬ì„± ì‹œ ì „ëŸ‰ ì²­ì‚°', '', '', ''],
        ['ìµœëŒ€ ë³´ìœ  ê¸°ê°„', '30ì¼ ê²½ê³¼ ì‹œ ìˆ˜ìµ/ì†ì‹¤ ë¬´ê´€ ì „ëŸ‰ ì²­ì‚°', '', '', ''],
        ['', '', '', '', ''],
        ['', '', '', '', ''],
        ['ğŸ“Œ ë³´ê³ ì„œ ì‘ì„±ì¼: 2026-02-14  |  ë¶„ì„ ì‹œìŠ¤í…œ: ìˆ˜ê¸‰ ë ˆì§ ìŠ¤ìºë„ˆ v3.1 (Stage 3 ì™„ë£Œ)', '', '', '', ''],
    ]

    # DataFrameìœ¼ë¡œ ë³€í™˜
    df_glossary = pd.DataFrame(glossary_data)

    # ì‹œíŠ¸ì— ì“°ê¸° (í—¤ë” ì—†ì´)
    df_glossary.to_excel(writer, sheet_name='0.ìš©ì–´ì„¤ëª…', index=False, header=False)

    # í¬ë§·íŒ…
    if HAS_OPENPYXL:
        ws = writer.sheets['0.ìš©ì–´ì„¤ëª…']
        from openpyxl.styles import Font, Alignment, PatternFill

        # ì—´ ë„ˆë¹„ ì„¤ì • (ê°€ë¡œë¡œ ë„“ê²Œ)
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40

        # ì „ì²´ ì…€ í¬ë§·íŒ…
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='ë§‘ì€ ê³ ë”•', size=9)
                cell.alignment = Alignment(vertical='top', wrap_text=False)  # ìë™ ì¤„ë°”ê¿ˆ OFF

                # ì œëª© (ì²« í–‰)
                if cell.row == 1:
                    cell.font = Font(name='ë§‘ì€ ê³ ë”•', size=14, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # ì„¹ì…˜ í—¤ë” (â–  í¬í•¨)
                if cell.value and 'â– ' in str(cell.value):
                    cell.font = Font(name='ë§‘ì€ ê³ ë”•', size=11, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # íŒ¨í„´/ì‹œê·¸ë„ ì´ë¦„ (ì´ëª¨ì§€ í¬í•¨)
                if cell.value and any(emoji in str(cell.value) for emoji in ['ğŸ”¥', 'ğŸ“ˆ', 'ğŸ”„', 'âœ…', 'âš¡', 'ğŸ¤', 'ğŸ“', 'ğŸ“Œ']):
                    cell.font = Font(name='ë§‘ì€ ê³ ë”•', size=10, bold=True, color='C00000')

        # ì²« í–‰ ë³‘í•© (ì œëª©)
        ws.merge_cells('A1:E1')

        # ë†’ì´ ì¡°ì •
        ws.row_dimensions[1].height = 25


def create_excel_report(csv_path: str, output_path: str, signal_bonus: int = 5):
    """
    ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„±

    Args:
        csv_path: ì…ë ¥ CSV íŒŒì¼ ê²½ë¡œ
        output_path: ì¶œë ¥ ì—‘ì…€ íŒŒì¼ ê²½ë¡œ
        signal_bonus: ì‹œê·¸ë„ 1ê°œë‹¹ ë³´ë„ˆìŠ¤ ì ìˆ˜
    """
    # CSV ì½ê¸° (ì¢…ëª©ì½”ë“œë¥¼ ë¬¸ìì—´ë¡œ ì½ì–´ ì•ì˜ 0 ë³´ì¡´)
    df = pd.read_csv(csv_path, encoding='utf-8-sig', dtype={'stock_code': str})

    # ì¢…ëª©ì½”ë“œ 6ìë¦¬ë¡œ íŒ¨ë”© í›„ 'A' ë¶™ì´ê¸° (ì—‘ì…€ì—ì„œ 0ìœ¼ë¡œ ì‹œì‘í•˜ëŠ” ì½”ë“œ ë³´í˜¸)
    df['stock_code'] = df['stock_code'].str.zfill(6)  # 6ìë¦¬ë¡œ íŒ¨ë”© (ì˜ˆ: 5930 â†’ 005930)
    df['stock_code'] = 'A' + df['stock_code']  # A ì ‘ë‘ì‚¬ ì¶”ê°€ (ì˜ˆ: 005930 â†’ A005930)

    # ì¢…í•© ì ìˆ˜ ê³„ì‚°
    df = calculate_combined_score(df, signal_bonus)

    print(f"[INFO] Loaded {len(df)} stocks")
    print(f"[INFO] Stock codes prefixed with 'A' to preserve leading zeros in Excel")
    print(f"[INFO] Combined score formula: ì›ë˜ì ìˆ˜ + (ì‹œê·¸ë„ Ã— {signal_bonus}ì )")

    # ì—‘ì…€ íŒŒì¼ ìƒì„±
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # ========================================
        # ì‹œíŠ¸ 0: ìš©ì–´ ì„¤ëª…
        # ========================================
        create_glossary_sheet(writer, signal_bonus)

        # ========================================
        # ì‹œíŠ¸ 1: ìµœì¢… ê²°ë¡  (ì¢…í•© ì¶”ì²œ ìˆœìœ„)
        # ========================================
        df_final = df.nlargest(20, 'combined_score')[[
            'stock_code', 'stock_name', 'sector', 'pattern',
            'score', 'signal_count', 'combined_score', 'signal_list',
            'entry_point', 'stop_loss'
        ]].copy()

        df_final.columns = ['ì¢…ëª©ì½”ë“œ', 'ì¢…ëª©ëª…', 'ì„¹í„°', 'íŒ¨í„´',
                           'ì ìˆ˜', 'ì‹œê·¸ë„', 'ì¢…í•©ì ìˆ˜', 'ì‹œê·¸ë„ë‚´ìš©',
                           'ì§„ì…ì „ëµ', 'ì†ì ˆ']
        df_final.insert(0, 'ìˆœìœ„', range(1, len(df_final) + 1))

        df_final.to_excel(writer, sheet_name='1.ìµœì¢…ê²°ë¡ ', index=False)
        format_excel_sheet(writer.sheets['1.ìµœì¢…ê²°ë¡ '], df_final,
                          f"ğŸ“Š ì¢…í•© ì¶”ì²œ ìˆœìœ„ TOP 20 (ì¢…í•©ì ìˆ˜ = ì ìˆ˜ + ì‹œê·¸ë„Ã—{signal_bonus})")

        # ========================================
        # ì‹œíŠ¸ 2: ì ìˆ˜ ìˆœìœ„ (ì¢…ëª©ì˜ ì§ˆ)
        # ========================================
        df_score = df.nlargest(30, 'score')[[
            'stock_code', 'stock_name', 'sector', 'pattern',
            'score', 'signal_count', 'recent', 'momentum', 'weighted', 'average'
        ]].copy()

        df_score.columns = ['ì¢…ëª©ì½”ë“œ', 'ì¢…ëª©ëª…', 'ì„¹í„°', 'íŒ¨í„´',
                           'ì ìˆ˜', 'ì‹œê·¸ë„', 'Recent', 'Momentum', 'Weighted', 'Average']
        df_score.insert(0, 'ìˆœìœ„', range(1, len(df_score) + 1))

        df_score.to_excel(writer, sheet_name='2.ì ìˆ˜ìˆœìœ„', index=False)
        format_excel_sheet(writer.sheets['2.ì ìˆ˜ìˆœìœ„'], df_score,
                          "ğŸ“ˆ ì ìˆ˜ ìˆœìœ„ TOP 30 (ì¢…ëª©ì˜ ì§ˆ)")

        # ========================================
        # ì‹œíŠ¸ 3: ì‹œê·¸ë„ ìˆœìœ„ (ì§„ì… íƒ€ì´ë°)
        # ========================================
        df_signal = df.sort_values(['signal_count', 'score'], ascending=[False, False]).head(30)[[
            'stock_code', 'stock_name', 'sector', 'pattern',
            'score', 'signal_count', 'signal_list', 'entry_point'
        ]].copy()

        df_signal.columns = ['ì¢…ëª©ì½”ë“œ', 'ì¢…ëª©ëª…', 'ì„¹í„°', 'íŒ¨í„´',
                            'ì ìˆ˜', 'ì‹œê·¸ë„', 'ì‹œê·¸ë„ë‚´ìš©', 'ì§„ì…ì „ëµ']
        df_signal.insert(0, 'ìˆœìœ„', range(1, len(df_signal) + 1))

        df_signal.to_excel(writer, sheet_name='3.ì‹œê·¸ë„ìˆœìœ„', index=False)
        format_excel_sheet(writer.sheets['3.ì‹œê·¸ë„ìˆœìœ„'], df_signal,
                          "ğŸš¨ ì‹œê·¸ë„ ìˆœìœ„ TOP 30 (ì§„ì… íƒ€ì´ë°)")

        # ========================================
        # ì‹œíŠ¸ 4: íŒ¨í„´ë³„ ìˆœìœ„
        # ========================================
        patterns = ['ëª¨ë©˜í…€í˜•', 'ì§€ì†í˜•', 'ì „í™˜í˜•']
        df_patterns_list = []

        for pattern in patterns:
            df_p = df[df['pattern'] == pattern].nlargest(10, 'combined_score')[[
                'stock_code', 'stock_name', 'sector',
                'score', 'signal_count', 'combined_score'
            ]].copy()

            if len(df_p) > 0:
                df_p.insert(0, 'íŒ¨í„´', pattern)
                df_p.insert(1, 'ìˆœìœ„', range(1, len(df_p) + 1))
                df_patterns_list.append(df_p)

        if df_patterns_list:
            df_patterns = pd.concat(df_patterns_list, ignore_index=True)
            df_patterns.columns = ['íŒ¨í„´', 'ìˆœìœ„', 'ì¢…ëª©ì½”ë“œ', 'ì¢…ëª©ëª…', 'ì„¹í„°',
                                  'ì ìˆ˜', 'ì‹œê·¸ë„', 'ì¢…í•©ì ìˆ˜']

            df_patterns.to_excel(writer, sheet_name='4.íŒ¨í„´ë³„ìˆœìœ„', index=False)
            format_excel_sheet(writer.sheets['4.íŒ¨í„´ë³„ìˆœìœ„'], df_patterns,
                              "ğŸ¯ íŒ¨í„´ë³„ TOP 10")

        # ========================================
        # ì‹œíŠ¸ 5: ì„¹í„°ë³„ ìƒìœ„
        # ========================================
        sectors = df.groupby('sector')['combined_score'].max().nlargest(20).index
        df_sectors_list = []

        for sector in sectors:
            df_s = df[df['sector'] == sector].nlargest(3, 'combined_score')[[
                'stock_code', 'stock_name', 'pattern',
                'score', 'signal_count', 'combined_score'
            ]].copy()

            df_s.insert(0, 'ì„¹í„°', sector)
            df_sectors_list.append(df_s)

        df_sectors = pd.concat(df_sectors_list, ignore_index=True)
        df_sectors.columns = ['ì„¹í„°', 'ì¢…ëª©ì½”ë“œ', 'ì¢…ëª©ëª…', 'íŒ¨í„´',
                             'ì ìˆ˜', 'ì‹œê·¸ë„', 'ì¢…í•©ì ìˆ˜']

        df_sectors.to_excel(writer, sheet_name='5.ì„¹í„°ë³„ìƒìœ„', index=False)
        format_excel_sheet(writer.sheets['5.ì„¹í„°ë³„ìƒìœ„'], df_sectors,
                          "ğŸ¢ ì„¹í„°ë³„ ìƒìœ„ ì¢…ëª© (ì„¹í„°ë‹¹ TOP 3)")

        # ========================================
        # ì‹œíŠ¸ 6: ì „ì²´ ë°ì´í„°
        # ========================================
        df_all = df.sort_values('combined_score', ascending=False).copy()
        df_all.insert(0, 'ìˆœìœ„', range(1, len(df_all) + 1))

        # ì»¬ëŸ¼ëª… í•œê¸€í™”
        column_mapping = {
            'stock_code': 'ì¢…ëª©ì½”ë“œ',
            'stock_name': 'ì¢…ëª©ëª…',
            'sector': 'ì„¹í„°',
            'pattern': 'íŒ¨í„´',
            'score': 'ì ìˆ˜',
            'recent': 'Recent',
            'momentum': 'Momentum',
            'weighted': 'Weighted',
            'average': 'Average',
            'signal_count': 'ì‹œê·¸ë„',
            'signal_list': 'ì‹œê·¸ë„ë‚´ìš©',
            'combined_score': 'ì¢…í•©ì ìˆ˜',
            'entry_point': 'ì§„ì…ì „ëµ',
            'stop_loss': 'ì†ì ˆ'
        }

        df_all = df_all.rename(columns=column_mapping)

        df_all.to_excel(writer, sheet_name='6.ì „ì²´ë°ì´í„°', index=False)
        format_excel_sheet(writer.sheets['6.ì „ì²´ë°ì´í„°'], df_all,
                          f"ğŸ“‹ ì „ì²´ ë°ì´í„° ({len(df_all)}ê°œ ì¢…ëª©)")

    print(f"âœ… Excel report saved: {output_path}")
    print(f"\nì‹œíŠ¸ êµ¬ì„±:")
    print(f"  0. ìš©ì–´ì„¤ëª… - íŒ¨í„´/ì‹œê·¸ë„/ë©”íŠ¸ë¦­ ì •ì˜")
    print(f"  1. ìµœì¢…ê²°ë¡  - ì¢…í•© ì¶”ì²œ ìˆœìœ„ TOP 20")
    print(f"  2. ì ìˆ˜ìˆœìœ„ - ì¢…ëª©ì˜ ì§ˆ TOP 30")
    print(f"  3. ì‹œê·¸ë„ìˆœìœ„ - ì§„ì… íƒ€ì´ë° TOP 30")
    print(f"  4. íŒ¨í„´ë³„ìˆœìœ„ - íŒ¨í„´ë³„ TOP 10")
    print(f"  5. ì„¹í„°ë³„ìƒìœ„ - ì„¹í„°ë³„ TOP 3")
    print(f"  6. ì „ì²´ë°ì´í„° - ì „ì²´ {len(df_all)}ê°œ ì¢…ëª©")

    # ========================================
    # CSV íŒŒì¼ë„ í•¨ê»˜ ì €ì¥
    # ========================================
    csv_output_path = output_path.replace('.xlsx', '.csv')
    df_all.to_csv(csv_output_path, index=False, encoding='utf-8-sig')
    print(f"\nâœ… CSV file saved: {csv_output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='ìˆ˜ê¸‰ ë ˆì§ ìŠ¤ìºë„ˆ - ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„±',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
ì˜ˆì‹œ:
  # ê¸°ë³¸ ì‹¤í–‰ (CSV â†’ Excel)
  python scripts/analysis/generate_excel_report.py

  # ì…ë ¥/ì¶œë ¥ íŒŒì¼ ì§€ì •
  python scripts/analysis/generate_excel_report.py --input output/regime_report.csv --output output/report.xlsx

  # ì‹œê·¸ë„ ë³´ë„ˆìŠ¤ ì ìˆ˜ ë³€ê²½ (ê¸°ë³¸ 5ì )
  python scripts/analysis/generate_excel_report.py --signal-bonus 10
        """
    )

    parser.add_argument(
        '--input', '-i',
        default='output/regime_report.csv',
        help='ì…ë ¥ CSV íŒŒì¼ (ê¸°ë³¸: output/regime_report.csv)'
    )

    parser.add_argument(
        '--output', '-o',
        default=None,
        help='ì¶œë ¥ ì—‘ì…€ íŒŒì¼ (ê¸°ë³¸: output/regime_report_YYYYMMDD_HHMMSS.xlsx)'
    )

    parser.add_argument(
        '--signal-bonus',
        type=int,
        default=5,
        help='ì‹œê·¸ë„ 1ê°œë‹¹ ë³´ë„ˆìŠ¤ ì ìˆ˜ (ê¸°ë³¸: 5)'
    )

    args = parser.parse_args()

    # ì…ë ¥ íŒŒì¼ í™•ì¸
    if not Path(args.input).exists():
        print(f"[ERROR] Input file not found: {args.input}")
        print("\në¨¼ì € regime_scanner.pyë¥¼ ì‹¤í–‰í•˜ì—¬ CSV íŒŒì¼ì„ ìƒì„±í•˜ì„¸ìš”:")
        print("  python scripts/analysis/regime_scanner.py --save-csv")
        return 1

    # ì¶œë ¥ íŒŒì¼ëª… ìƒì„±
    if args.output is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        args.output = f'output/regime_report_{timestamp}.xlsx'

    # ë””ë ‰í† ë¦¬ ìƒì„±
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)

    # ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„±
    print("="*80)
    print("ğŸ“Š ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„±")
    print("="*80)
    print(f"ì…ë ¥: {args.input}")
    print(f"ì¶œë ¥: {args.output}")
    print(f"ì‹œê·¸ë„ ë³´ë„ˆìŠ¤: {args.signal_bonus}ì ")
    print()

    create_excel_report(args.input, args.output, args.signal_bonus)

    print()
    print("="*80)
    print("âœ… ì™„ë£Œ!")
    print("="*80)

    return 0


if __name__ == '__main__':
    exit(main())
